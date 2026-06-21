"""
excel_exporter.py
Builds the downloadable Product_Review_Analysis_Report.xlsx — a clean,
business-ready sheet of every review plus a summary sheet.
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ai_analyzer import AnalysisResult
from utils import basic_keyword_sentiment, get_logger

logger = get_logger(__name__)

HEADER_FILL = PatternFill("solid", start_color="1F2A44", end_color="1F2A44")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F2A44")
LABEL_FONT = Font(name="Calibri", bold=True, size=10, color="1F2A44")

SENTIMENT_FILL = {
    "Positive": PatternFill("solid", start_color="D9F2D9", end_color="D9F2D9"),
    "Neutral": PatternFill("solid", start_color="FFF6CC", end_color="FFF6CC"),
    "Negative": PatternFill("solid", start_color="FBE0E0", end_color="FBE0E0"),
}


class ExcelExporter:

    @staticmethod
    def build_report(df: pd.DataFrame, product_name: str, analysis: AnalysisResult) -> bytes:
        """Returns the .xlsx file as bytes, ready for a Streamlit download_button."""
        wb = Workbook()

        ExcelExporter._build_summary_sheet(wb, product_name, analysis, total_reviews=len(df))
        ExcelExporter._build_reviews_sheet(wb, df, product_name)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ---------- summary sheet ----------

    @staticmethod
    def _build_summary_sheet(wb: Workbook, product_name: str, analysis: AnalysisResult, total_reviews: int):
        ws = wb.active
        ws.title = "Summary"

        ws["A1"] = "AI Product Review Intelligence Report"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:B1")

        ws["A2"] = product_name
        ws["A2"].font = Font(name="Calibri", italic=True, size=11)
        ws.merge_cells("A2:B2")

        rows = [
            ("Total Reviews Analyzed", total_reviews),
            ("Overall Product Score", f"{analysis.overall_score} / 10"),
            ("Final Verdict", analysis.verdict),
            ("Positive Sentiment", f"{analysis.positive_pct}%"),
            ("Neutral Sentiment", f"{analysis.neutral_pct}%"),
            ("Negative Sentiment", f"{analysis.negative_pct}%"),
            ("AI Provider Used", analysis.ai_provider_used),
        ]

        r = 4
        for label, value in rows:
            ws.cell(row=r, column=1, value=label).font = LABEL_FONT
            ws.cell(row=r, column=2, value=value).font = BODY_FONT
            r += 1

        r += 1
        ws.cell(row=r, column=1, value="What Customers Like").font = LABEL_FONT
        r += 1
        for item in (analysis.strengths or ["No standout strengths identified."]):
            ws.cell(row=r, column=1, value=f"• {item}").font = BODY_FONT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            r += 1

        r += 1
        ws.cell(row=r, column=1, value="Common Complaints").font = LABEL_FONT
        r += 1
        for item in (analysis.complaints or ["No major complaints identified."]):
            ws.cell(row=r, column=1, value=f"• {item}").font = BODY_FONT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            r += 1

        r += 1
        ws.cell(row=r, column=1, value="Buy Recommendation").font = LABEL_FONT
        r += 1
        ws.cell(row=r, column=1, value=analysis.recommendation or "Not available.").font = BODY_FONT
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=4)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22

    # ---------- reviews sheet ----------

    @staticmethod
    def _build_reviews_sheet(wb: Workbook, df: pd.DataFrame, product_name: str):
        ws = wb.create_sheet("Reviews")

        headers = ["Product Name", "Reviewer Name", "Rating", "Full Review Text", "Review Date", "Sentiment Label"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        for i, (_, row) in enumerate(df.iterrows(), start=2):
            sentiment = basic_keyword_sentiment(row.get("review_text", ""), row.get("rating"))

            values = [
                product_name,
                row.get("reviewer_name", "Anonymous"),
                row.get("rating", ""),
                row.get("review_text", ""),
                row.get("review_date", ""),
                sentiment,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.font = BODY_FONT
                cell.alignment = Alignment(wrap_text=(col == 4), vertical="top")
                if col == 6:
                    cell.fill = SENTIMENT_FILL.get(sentiment, PatternFill())

        widths = {1: 22, 2: 18, 3: 9, 4: 70, 5: 14, 6: 14}
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        for i in range(2, len(df) + 2):
            ws.row_dimensions[i].height = 45
